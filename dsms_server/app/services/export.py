"""
数据导出服务模块
支持JSON、CSV、Excel格式导出
"""

from typing import Optional, List, Dict, Any
from uuid import UUID
from datetime import datetime
import json
import csv
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from sqlmodel import select

from app.config.database import db_config
from app.models.models import (
    ProductionLine,
    DeviceApproval,
    Device,
    DeviceStatusHistory,
    DefectType,
    DetectionRecord,
    DefectDetail,
    ReviewTask,
    Role,
    Department,
    Title,
    User,
    UserOperationLog,
    UserMessage,
    SystemMessage,
    Announcement,
    AnnouncementReader,
)


class ExportService:
    """数据导出服务"""

    MODEL_TABLE_MAPPING = {
        "production_lines": ProductionLine,
        "device_approvals": DeviceApproval,
        "devices": Device,
        "device_status_history": DeviceStatusHistory,
        "defect_types": DefectType,
        "detection_records": DetectionRecord,
        "defect_details": DefectDetail,
        "review_tasks": ReviewTask,
        "roles": Role,
        "departments": Department,
        "titles": Title,
        "users": User,
        "user_operation_logs": UserOperationLog,
        "user_messages": UserMessage,
        "system_messages": SystemMessage,
        "announcements": Announcement,
        "announcement_readers": AnnouncementReader,
    }

    @staticmethod
    def get_all_table_names() -> List[str]:
        """获取所有可导出的表名"""
        return list(ExportService.MODEL_TABLE_MAPPING.keys())

    @staticmethod
    def get_table_data(table_name: str) -> List[Dict[str, Any]]:
        """获取指定表的所有数据"""
        model_class = ExportService.MODEL_TABLE_MAPPING.get(table_name)
        if not model_class:
            return []

        with db_config.get_session() as session:
            results = session.execute(select(model_class))
            records = results.scalars().all()

            data = []
            for record in records:
                record_dict = {}
                for key, value in record.__dict__.items():
                    if not key.startswith("_"):
                        if isinstance(value, UUID):
                            record_dict[key] = str(value)
                        elif isinstance(value, datetime):
                            record_dict[key] = value.isoformat()
                        else:
                            record_dict[key] = value
                data.append(record_dict)
            return data

    @staticmethod
    def get_all_data() -> Dict[str, List[Dict[str, Any]]]:
        """获取所有表的数据"""
        all_data = {}
        for table_name in ExportService.MODEL_TABLE_MAPPING.keys():
            all_data[table_name] = ExportService.get_table_data(table_name)
        return all_data

    @staticmethod
    def export_to_json(data: Dict[str, List[Dict[str, Any]]], pretty: bool = True) -> str:
        """导出为JSON格式"""
        if pretty:
            return json.dumps(data, indent=2, ensure_ascii=False, default=str)
        return json.dumps(data, ensure_ascii=False, default=str)

    @staticmethod
    def export_table_to_csv(table_name: str, data: List[Dict[str, Any]]) -> str:
        """导出指定表数据为CSV格式"""
        if not data:
            return ""

        output = io.StringIO()
        headers = list(data[0].keys())
        writer = csv.DictWriter(output, fieldnames=headers)

        writer.writeheader()
        for row in data:
            # 处理二进制数据
            processed_row = {}
            for key, value in row.items():
                if isinstance(value, bytes):
                    processed_row[key] = "(二进制数据，已跳过)"
                else:
                    processed_row[key] = value
            writer.writerow(processed_row)

        return output.getvalue()

    @staticmethod
    def export_all_to_csv(data: Dict[str, List[Dict[str, Any]]]) -> Dict[str, str]:
        """导出所有表数据为CSV格式（每个表一个CSV）"""
        result = {}
        for table_name, table_data in data.items():
            result[table_name] = ExportService.export_table_to_csv(table_name, table_data)
        return result

    @staticmethod
    def export_to_excel(data: Dict[str, List[Dict[str, Any]]]) -> bytes:
        """导出为Excel格式（.xlsx）"""
        wb = Workbook()
        wb.remove(wb.active)

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        for table_name, table_data in data.items():
            ws = wb.create_sheet(title=table_name[:31])

            if not table_data:
                continue

            headers = list(table_data[0].keys())
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            for row_idx, row_data in enumerate(table_data, 2):
                for col_idx, header in enumerate(headers, 1):
                    value = row_data.get(header, "")
                    if isinstance(value, datetime):
                        value = value.isoformat() if hasattr(value, 'isoformat') else str(value)
                    elif isinstance(value, (list, dict)):
                        value = json.dumps(value, default=str) if value else ""
                    elif isinstance(value, bytes):
                        value = "(二进制数据，已跳过)"
                    ws.cell(row=row_idx, column=col_idx, value=value)

            for col_idx in range(1, len(headers) + 1):
                ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 20

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    @staticmethod
    def export_table_to_excel(table_name: str, data: List[Dict[str, Any]]) -> bytes:
        """导出指定表数据为Excel格式"""
        wb = Workbook()
        wb.remove(wb.active)

        ws = wb.create_sheet(title=table_name[:31])

        if not data:
            return b""

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        headers = list(data[0].keys())
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        for row_idx, row_data in enumerate(data, 2):
            for col_idx, header in enumerate(headers, 1):
                value = row_data.get(header, "")
                if isinstance(value, datetime):
                    value = value.isoformat() if hasattr(value, 'isoformat') else str(value)
                elif isinstance(value, (list, dict)):
                    value = json.dumps(value, default=str) if value else ""
                elif isinstance(value, bytes):
                    value = "(二进制数据，已跳过)"
                ws.cell(row=row_idx, column=col_idx, value=value)

        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 20

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()


export_service = ExportService()